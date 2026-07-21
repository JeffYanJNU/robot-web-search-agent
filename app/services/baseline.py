from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from threading import Lock
from urllib.parse import urlparse

from openpyxl import load_workbook


def normalize_company_name(name: str) -> str:
    value = str(name or "").casefold().strip()
    value = re.sub(r"[\s\-_,.，。()（）·]+", "", value)
    for suffix in (
        "股份有限公司", "有限责任公司", "有限公司", "集团公司", "集团",
        "incorporated", "corporation", "limited", "company", "inc", "corp", "ltd", "co",
    ):
        value = value.removesuffix(suffix)
    return value


def normalize_code(value: object) -> str:
    return re.sub(r"[^0-9A-Z]", "", str(value or "").upper())


def normalize_domain(value: object) -> str:
    text = str(value or "").strip()
    if not text or text == "-":
        return ""
    parsed = urlparse(text if "://" in text else f"https://{text}")
    return (parsed.hostname or "").lower().removeprefix("www.")


@dataclass
class BaselineCompany:
    canonical_name: str
    credit_code: str = ""
    names: set[str] = field(default_factory=set)
    domains: set[str] = field(default_factory=set)
    evidence_text: str = ""


@dataclass(frozen=True)
class BaselineMatch:
    company: BaselineCompany
    matched_by: str


class BaselineRegistry:
    def __init__(self, path: str):
        self.path = Path(path).expanduser().resolve()
        if not self.path.is_file():
            raise FileNotFoundError(f"基线工作簿不存在：{self.path}")
        self.by_name: dict[str, BaselineCompany] = {}
        self.by_code: dict[str, BaselineCompany] = {}
        self.by_domain: dict[str, BaselineCompany] = {}
        self._load()

    def _load(self) -> None:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, [])]
            positions = {header: index for index, header in enumerate(headers)}
            name_col = positions.get("企查查匹配企业名称", positions.get("企业名称"))
            if name_col is None:
                continue
            code_col = positions.get("统一社会信用代码")
            alias_cols = [positions[name] for name in ("曾用名", "英文名") if name in positions]
            website_col = positions.get("官网")
            for row in rows:
                name = str(row[name_col] or "").strip() if name_col < len(row) else ""
                if not name:
                    continue
                code = normalize_code(row[code_col]) if code_col is not None and code_col < len(row) else ""
                company = self.by_code.get(code) if code else None
                company = company or self.by_name.get(normalize_company_name(name))
                if company is None:
                    company = BaselineCompany(canonical_name=name, credit_code=code)
                raw_names = [name]
                for index in alias_cols:
                    if index < len(row) and row[index] not in (None, "", "-"):
                        raw_names.extend(re.split(r"[;；]", str(row[index])))
                for raw_name in raw_names:
                    normalized = normalize_company_name(raw_name)
                    if normalized:
                        company.names.add(str(raw_name).strip())
                        self.by_name[normalized] = company
                if code:
                    company.credit_code = company.credit_code or code
                    self.by_code[code] = company
                if website_col is not None and website_col < len(row):
                    domain = normalize_domain(row[website_col])
                    if domain:
                        company.domains.add(domain)
                        self.by_domain[domain] = company
                text_parts = [str(value) for value in row if value not in (None, "", "-")]
                company.evidence_text = (company.evidence_text + " " + " ".join(text_parts))[-30000:]
        workbook.close()

    def match(self, names: list[str], credit_code: str = "", website: str = "") -> BaselineMatch | None:
        code = normalize_code(credit_code)
        if code and code in self.by_code:
            return BaselineMatch(self.by_code[code], "统一社会信用代码")
        domain = normalize_domain(website)
        if domain and domain in self.by_domain:
            return BaselineMatch(self.by_domain[domain], "官网域名")
        for name in names:
            normalized = normalize_company_name(name)
            if normalized and normalized in self.by_name:
                return BaselineMatch(self.by_name[normalized], "企业名称")
        return None


_cache_lock = Lock()
_cache: tuple[str, int, BaselineRegistry] | None = None


def get_baseline_registry(path: str) -> BaselineRegistry:
    global _cache
    resolved = str(Path(path).expanduser().resolve())
    mtime = Path(resolved).stat().st_mtime_ns
    with _cache_lock:
        if _cache is None or _cache[:2] != (resolved, mtime):
            _cache = (resolved, mtime, BaselineRegistry(resolved))
        return _cache[2]
