from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import ProductCompanyRelation, ProductSource, RobotCompany, RobotProduct
from app.schemas import RunResult
from app.services.product_inventory_matcher import compare_product_names_from_workbook
from app.services.product_rules import (
    PRODUCT_EVENT_TYPES,
    STRONG_RELATION_TYPES,
    is_same_product_identity,
    normalize_product_name,
)


NAVY = "172554"
BLUE = "3157E7"
PALE_GREEN = "DCFCE7"
PALE_AMBER = "FEF3C7"
PALE_RED = "FEE2E2"
WHITE = "FFFFFF"
TEXT = "172033"
GRID = "DDE3EC"

MAIN_HEADERS = [
    "A｜机器人产品名称",
    "B｜关联企业（简称 / 工商全称 / 统一社会信用代码）",
    "C｜产品是否存在及依据",
    "D｜产品与企业是否对应及依据",
    "E｜与已有产品名称相似度",
    "F｜相似度说明",
]

DETAIL_HEADERS = [
    "产品名称",
    "企业简称",
    "企业全称",
    "企业全称来源",
    "检索热度（来源页数）",
    "产品真实性评分",
    "新产品置信度",
    "关系置信度",
    "产品核验状态",
    "关系核验状态",
    "产品类别",
    "产品型号",
    "产品系列",
    "发布状态",
    "发布日期",
    "关系类型",
    "是否主要关系",
    "全部产品来源",
    "产品真实性证据",
    "产品—企业关系证据",
    "产品评分依据",
    "关系评分依据",
    "产品核验说明",
    "关系核验说明",
    "企业官网",
    "统一社会信用代码",
    "任务模式",
    "任务统计",
    "生成时间",
]

QCC_DIAGNOSTIC_HEADERS = [
    "查询企业名称",
    "候选企业名称",
    "统一社会信用代码",
    "名称相似度",
    "是否采用",
    "采用 / 拒绝原因",
]


def _json_list(value: str | None) -> list[dict[str, Any]]:
    try:
        parsed = json.loads(value or "[]")
    except (json.JSONDecodeError, TypeError):
        return []
    return [item for item in parsed if isinstance(item, dict)] if isinstance(parsed, list) else []


def _unique(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for raw in values:
        value = str(raw or "").strip()
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return output


def _limit_text(value: str, limit: int = 32000) -> str:
    return value if len(value) <= limit else value[: limit - 12] + "\n……内容已截断"


LEGAL_SUFFIX = re.compile(
    r"(?:股份有限公司|有限责任公司|集团有限公司|有限公司|公司)$",
    re.IGNORECASE,
)


def resolve_company_names(company: RobotCompany) -> tuple[str, str, str]:
    candidates = _unique(
        [
            company.baseline_company_name,
            company.canonical_name,
            company.original_name,
            company.chinese_name,
        ]
    )
    full_name = next((name for name in candidates if LEGAL_SUFFIX.search(name)), "")
    if company.baseline_company_name and full_name == company.baseline_company_name:
        full_source = "Excel 企业基线"
    elif company.unified_social_credit_code and full_name:
        full_source = "企业工商 API 核验"
    elif full_name:
        full_source = "现有企业主体数据"
    else:
        full_name = company.canonical_name
        full_source = "候选名称，待天眼查/工商接口核验"
    short_candidates = [
        name for name in candidates if name != full_name and not LEGAL_SUFFIX.search(name)
    ]
    short_name = min(short_candidates, key=len) if short_candidates else LEGAL_SUFFIX.sub("", full_name)
    return short_name or company.canonical_name, full_name, full_source


class _RelationExportView:
    """Best relation plus evidence collected from equivalent legacy rows."""

    def __init__(self, relations: list[ProductCompanyRelation]):
        self.primary = max(
            relations,
            key=lambda item: (
                item.is_primary,
                item.verification_status == "verified",
                item.relation_score,
                len(item.evidence_json or ""),
            ),
        )
        evidence: list[dict[str, Any]] = []
        seen: set[tuple[str, str, str, str]] = set()
        for relation in relations:
            for item in _json_list(relation.evidence_json):
                key = (
                    str(item.get("source_url") or ""),
                    str(item.get("quote") or ""),
                    str(item.get("company_name") or ""),
                    str(item.get("relation_type") or ""),
                )
                if key not in seen:
                    seen.add(key)
                    evidence.append(item)
        self.company_id = self.primary.company_id
        self.relation_type = self.primary.relation_type
        self.relation_score = max(item.relation_score for item in relations)
        self.verification_status = (
            "verified"
            if any(item.verification_status == "verified" for item in relations)
            else self.primary.verification_status
        )
        self.verification_reason = "；".join(
            _unique(item.verification_reason for item in relations)
        )
        self.evidence_json = json.dumps(evidence, ensure_ascii=False)
        self.is_primary = any(item.is_primary for item in relations)

    def __getattr__(self, name: str) -> Any:
        return getattr(self.primary, name)


class _ProductExportView:
    """Read-only merged view used to suppress legacy duplicate products."""

    def __init__(
        self,
        products: list[RobotProduct],
        allowed_company_ids: set[int] | None = None,
    ):
        self.products = products
        self.representative = max(
            products,
            key=lambda item: (
                len({source.source_url for source in item.sources}),
                item.authenticity_score,
                item.novelty_score,
                len(item.canonical_name or ""),
            ),
        )
        source_by_url: dict[str, ProductSource] = {}
        for product in products:
            for source in product.sources:
                key = source.canonical_url or source.source_url
                current = source_by_url.get(key)
                if current is None or len(source.evidence_json or "") > len(
                    current.evidence_json or ""
                ):
                    source_by_url[key] = source
        self.sources = list(source_by_url.values())

        relation_by_key: dict[tuple[int, str], ProductCompanyRelation] = {}
        for product in products:
            for relation in product.company_relations:
                if (
                    allowed_company_ids is not None
                    and relation.company_id not in allowed_company_ids
                ):
                    continue
                key = (relation.company_id, relation.relation_type)
                current = relation_by_key.get(key)
                if current is None or (
                    relation.relation_score,
                    len(relation.evidence_json or ""),
                ) > (
                    current.relation_score,
                    len(current.evidence_json or ""),
                ):
                    relation_by_key[key] = relation
        relations = list(relation_by_key.values())
        self.company_relations = [_RelationExportView(relations)] if relations else []

    def __getattr__(self, name: str) -> Any:
        return getattr(self.representative, name)


def _group_products_for_export(
    products: list[RobotProduct],
    allowed_company_ids: set[int] | None = None,
) -> list[_ProductExportView]:
    groups: list[list[RobotProduct]] = []
    normalized_by_id = {
        product.product_id: normalize_product_name(
            product.canonical_name,
            product.model_number,
            product.series_name,
        )
        for product in products
    }
    for product in products:
        normalized = normalized_by_id[product.product_id]
        for group in groups:
            if any(
                is_same_product_identity(
                    normalized,
                    normalized_by_id[member.product_id],
                )
                for member in group
            ):
                group.append(product)
                break
        else:
            groups.append([product])
    views = [
        _ProductExportView(group, allowed_company_ids=allowed_company_ids)
        for group in groups
    ]
    if allowed_company_ids is not None:
        views = [view for view in views if view.company_relations]
    return views


def _product_evidence(product: RobotProduct) -> list[dict[str, Any]]:
    return [
        {
            **item,
            "source_url": item.get("source_url") or source.source_url,
            "source_type": item.get("source_type") or source.source_type,
        }
        for source in product.sources
        for item in _json_list(source.evidence_json)
    ]


def _evidence_lines(evidence: list[dict[str, Any]], limit: int = 8) -> str:
    lines: list[str] = []
    for item in evidence:
        quote = str(item.get("quote") or "").strip()
        url = str(item.get("source_url") or "").strip()
        evidence_type = str(item.get("evidence_type") or "证据").strip()
        if not quote and not url:
            continue
        line = f"[{evidence_type}] {quote}" if quote else f"[{evidence_type}]"
        if url:
            line += f"\n来源：{url}"
        lines.append(line)
    return _limit_text("\n\n".join(_unique(lines)[:limit]))


def _product_truth_text(product: RobotProduct, evidence: list[dict[str, Any]]) -> str:
    verdict = {
        "verified": "是，产品存在且已核验",
        "needs_review": "发现产品线索，仍待补充核验",
        "rejected": "否或证据不足，已排除",
    }.get(product.verification_status, product.verification_status or "待核验")
    basis = _evidence_lines(evidence, limit=5)
    return _limit_text("\n".join(part for part in [verdict, product.verification_reason, basis] if part))


def _relation_evidence(relation: ProductCompanyRelation | None) -> list[dict[str, Any]]:
    return _json_list(relation.evidence_json) if relation else []


def _relation_truth_text(relation: ProductCompanyRelation | None) -> str:
    if relation is None:
        return "未发现明确的产品—企业对应关系证据"
    if relation.relation_type in STRONG_RELATION_TYPES:
        verdict = (
            "是，产品与该企业存在明确归属、研发或制造关系"
            if relation.verification_status == "verified"
            else "存在较强对应关系线索，但仍待核验"
        )
    else:
        verdict = f"存在“{relation.relation_type}”关系，但不等同于产品归属该企业"
    basis = _evidence_lines(_relation_evidence(relation), limit=5)
    return _limit_text("\n".join(part for part in [verdict, relation.verification_reason, basis] if part))


def _product_score_basis(product: RobotProduct, lookback_days: int) -> str:
    evidence = _product_evidence(product)
    clusters = {source.claim_fingerprint or source.content_hash for source in product.sources}
    all_urls = _unique(source.source_url for source in product.sources)
    identity_urls = _unique(
        item.get("source_url", "")
        for item in evidence
        if item.get("evidence_type") in {"product_identity", "official_product_page"}
        or item.get("value")
    )
    event_urls = _unique(
        item.get("source_url", "")
        for item in evidence
        if item.get("evidence_type") in PRODUCT_EVENT_TYPES
    )
    dated_urls = _unique(
        item.get("source_url", "") for item in evidence if item.get("evidence_date")
    )
    trusted_urls = _unique(
        source.source_url
        for source in product.sources
        if source.source_type in {"official", "authority"}
    )
    commercial_urls = _unique(
        item.get("source_url", "")
        for item in evidence
        if item.get("evidence_type")
        in {"technical_spec", "mass_production", "delivery", "order"}
    )
    recent = bool(
        product.launch_date
        and product.launch_date >= date.today() - timedelta(days=lookback_days)
    )
    novelty_urls = _unique(
        item.get("source_url", "")
        for item in evidence
        if re.search(r"新品|首款|新一代|新型号|首次发布|全新", str(item.get("quote") or ""))
    )
    rules = [
        ("真实性", "产品名称或型号有原文证据", 25, bool(identity_urls), identity_urls),
        ("真实性", "存在发布、亮相、量产或交付事件", 20, bool(event_urls), event_urls),
        ("真实性", "存在明确事件日期", 10, bool(dated_urls or product.launch_date), dated_urls),
        ("真实性", "存在官方、政府或权威来源", 20, bool(trusted_urls), trusted_urls),
        ("真实性", "至少两个非转载事实来源", 20, len(clusters) >= 2, all_urls),
        ("真实性", "存在参数、量产、交付或订单证据", 5, bool(commercial_urls), commercial_urls),
        ("新产品", "发布时间在任务回溯期内", 35, recent, event_urls),
        (
            "新产品",
            "历史产品库没有相同产品",
            25,
            product.addition_type not in {"historical_product", "upgrade"},
            [],
        ),
        ("新产品", "来源明确称新品、首款或新一代", 15, bool(novelty_urls), novelty_urls),
        ("新产品", "至少两个非转载事实来源", 15, len(clusters) >= 2, all_urls),
        ("新产品", "型号或版本明确", 10, bool(product.model_number), all_urls),
    ]
    lines = [
        f"最终分数：真实性 {product.authenticity_score}；新产品置信度 {product.novelty_score}"
    ]
    for score_type, criterion, weight, met, urls in rules:
        status = "满足" if met else "不满足"
        source_text = "；".join(urls) if urls else "—"
        lines.append(
            f"[{score_type}] {criterion}｜{status}｜权重 {weight}｜得分 {weight if met else 0}｜来源：{source_text}"
        )
    return _limit_text("\n".join(lines))


def _relation_score_basis(
    product: RobotProduct,
    relation: ProductCompanyRelation | None,
    company: RobotCompany | None,
) -> str:
    if relation is None or company is None:
        return "无关系记录，关系置信度为 0"
    evidence = _relation_evidence(relation)
    urls = _unique(item.get("source_url", "") for item in evidence)
    source_by_url = {source.source_url: source for source in product.sources}
    relation_sources = [source_by_url[url] for url in urls if url in source_by_url]
    clusters = {source.claim_fingerprint or source.content_hash for source in relation_sources}
    official_urls = _unique(
        source.source_url for source in relation_sources if source.source_type == "official"
    )
    rules = [
        ("企业、产品和关系动作有明确原文", 50, bool(evidence), urls),
        ("关系证据来自企业官网", 20, bool(official_urls), official_urls),
        ("至少两个独立来源确认关系", 20, len(clusters) >= 2, urls),
        (
            "企业官网域名或主体身份已确认",
            10,
            bool(company.official_domain or company.unified_social_credit_code),
            [],
        ),
    ]
    lines = [f"最终分数：关系置信度 {relation.relation_score}"]
    for criterion, weight, met, rule_urls in rules:
        source_text = "；".join(rule_urls) if rule_urls else "—"
        lines.append(
            f"{criterion}｜{'满足' if met else '不满足'}｜权重 {weight}｜得分 {weight if met else 0}｜来源：{source_text}"
        )
    return _limit_text("\n".join(lines))


def _source_details(product: RobotProduct) -> str:
    lines = []
    for source in product.sources:
        published = source.published_at.isoformat() if source.published_at else "日期未知"
        lines.append(
            f"{source.source_title or '未命名来源'}｜{source.source_type or 'unknown'}｜{published}\n{source.source_url}"
        )
    return _limit_text("\n\n".join(_unique(lines)))


def _add_table_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    table_name: str,
    row_height: int = 54,
):
    worksheet = workbook.create_sheet(name)
    worksheet.sheet_view.showGridLines = False
    last_column = get_column_letter(len(headers))
    worksheet.merge_cells(f"A1:{last_column}1")
    title_cell = worksheet["A1"]
    title_cell.value = title
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(color=WHITE, bold=True, size=16)
    title_cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 30
    for index, header in enumerate(headers, 1):
        cell = worksheet.cell(2, index, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 38
    thin_bottom = Side(style="thin", color=GRID)
    for row_index, row in enumerate(rows, 3):
        for column_index, value in enumerate(row, 1):
            cell = worksheet.cell(row_index, column_index, value)
            cell.font = Font(color=TEXT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin_bottom)
        worksheet.row_dimensions[row_index].height = row_height
    worksheet.freeze_panes = "A3"
    if rows:
        table = Table(displayName=table_name, ref=f"A2:{last_column}{len(rows) + 2}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(width, 65)
    if not rows:
        worksheet.auto_filter.ref = f"A2:{last_column}2"
    return worksheet


def export_run_results(
    db: Session,
    result: RunResult,
    *,
    pipeline_mode: str,
    lookback_days: int,
    output_dir: str,
    run_id: str = "",
    inventory_workbook_path: str | None = None,
) -> Path:
    product_ids = list(dict.fromkeys(result.product_ids))
    company_ids = list(dict.fromkeys(result.company_ids))
    products = (
        list(
            db.scalars(
                select(RobotProduct)
                .options(
                    selectinload(RobotProduct.sources),
                    selectinload(RobotProduct.company_relations),
                )
                .where(RobotProduct.product_id.in_(product_ids))
            ).unique()
        )
        if product_ids
        else []
    )
    products.sort(
        key=lambda item: (
            len({source.source_url for source in item.sources}),
            item.authenticity_score,
            item.novelty_score,
        ),
        reverse=True,
    )
    related_company_ids = {
        relation.company_id
        for product in products
        for relation in product.company_relations
    }
    all_company_ids = set(company_ids) | related_company_ids
    companies = (
        list(
            db.scalars(
                select(RobotCompany).where(RobotCompany.company_id.in_(all_company_ids))
            ).unique()
        )
        if all_company_ids
        else []
    )
    company_by_id = {company.company_id: company for company in companies}
    mainland_company_ids = {
        company.company_id
        for company in companies
        if company.region_type == "mainland_china"
    }
    products = _group_products_for_export(
        products,
        allowed_company_ids=mainland_company_ids,
    )

    main_rows: list[list[Any]] = []
    detail_rows: list[list[Any]] = []
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    task_stats = (
        f"查询 {result.queries}；搜索结果 {result.results}；抓取网页 {result.fetched}；"
        f"原始产品候选 {result.raw_product_candidates}；自动修复 {result.repaired_product_candidates}；"
        f"有效产品候选 {result.product_candidates}；阶段入库 {result.products_staged}；"
        f"新增产品 {result.products_created}；更新产品 {result.products_updated}；"
        f"新增关系 {result.relations_created}；已核验关系 {result.relations_verified}；"
        f"企业工商 API（{result.qcc_provider or '未配置'}）"
        f" {result.qcc_api_calls}/{result.qcc_api_limit}；"
        f"返回候选 {result.qcc_candidates}；工商主体命中 {result.qcc_matches}；"
        f"未匹配查询 {result.qcc_unmatched}；接口错误 {result.qcc_api_errors}；"
        f"错误 {len(result.errors)}"
    )
    qcc_diagnostic_rows = [
        [
            str(item.get("query_name") or ""),
            str(item.get("candidate_name") or ""),
            str(item.get("credit_code") or ""),
            float(item.get("similarity") or 0) / 100.0,
            "是" if item.get("accepted") else "否",
            str(item.get("reason") or ""),
        ]
        for item in result.qcc_match_diagnostics
    ]

    for product in products:
        product_evidence = _product_evidence(product)
        sources = _unique(source.source_url for source in product.sources)
        relations = sorted(
            product.company_relations,
            key=lambda item: (item.is_primary, item.relation_score),
            reverse=True,
        )
        for relation in relations or [None]:
            company = company_by_id.get(relation.company_id) if relation else None
            if company:
                short_name, full_name, full_source = resolve_company_names(company)
            else:
                short_name, full_name, full_source = "未识别", "待核验", "未识别到对应企业"

            credit_code = (
                company.unified_social_credit_code
                if company and company.unified_social_credit_code
                else "待工商核验"
            )
            verified_full_name = (
                full_name
                if company and (
                    company.unified_social_credit_code
                    or full_source in {"Excel 企业基线", "企业工商 API 核验"}
                )
                else f"{full_name}（待工商核验）"
            )
            company_display = (
                f"简称：{short_name}\n"
                f"工商全称：{verified_full_name}\n"
                f"统一社会信用代码：{credit_code}"
            )
            main_rows.append(
                [
                    product.canonical_name,
                    company_display,
                    _product_truth_text(product, product_evidence),
                    _relation_truth_text(relation),
                ]
            )

            detail_rows.append(
                [
                    product.canonical_name,
                    short_name,
                    full_name,
                    full_source,
                    len(sources),
                    product.authenticity_score,
                    product.novelty_score,
                    relation.relation_score if relation else 0,
                    product.verification_status,
                    relation.verification_status if relation else "无关系",
                    product.robot_category,
                    product.model_number,
                    product.series_name,
                    product.launch_status,
                    product.launch_date.isoformat() if product.launch_date else "",
                    relation.relation_type if relation else "",
                    "是" if relation and relation.is_primary else "否",
                    _source_details(product),
                    _evidence_lines(product_evidence),
                    _evidence_lines(_relation_evidence(relation)),
                    _product_score_basis(product, lookback_days),
                    _relation_score_basis(product, relation, company),
                    product.verification_reason,
                    relation.verification_reason if relation else "",
                    company.official_website if company else "",
                    company.unified_social_credit_code if company else "",
                    pipeline_mode,
                    task_stats,
                    generated_at,
                ]
            )

    if main_rows:
        if inventory_workbook_path:
            name_matches = compare_product_names_from_workbook(
                (str(row[0]) for row in main_rows),
                inventory_workbook_path,
            )
            for row, name_match in zip(main_rows, name_matches, strict=True):
                row.extend([name_match.score / 100.0, name_match.explanation])
        else:
            for row in main_rows:
                row.extend([0.0, "任务开始前未指定已有产品库存文件，未执行名称相似度对比"])

    workbook = Workbook()
    workbook.remove(workbook.active)
    main = _add_table_sheet(
        workbook,
        "结果主表",
        "高热度机器人产品、关联企业与真实性核验结果",
        MAIN_HEADERS,
        main_rows,
        [26, 42, 65, 65, 20, 65],
        "ProductResultsTable",
        row_height=96,
    )
    detail = _add_table_sheet(
        workbook,
        "详细信息",
        "来源、评分依据、核验状态与任务信息",
        DETAIL_HEADERS,
        detail_rows,
        [24, 18, 32, 28, 16, 16, 16, 16, 18, 18, 18, 16, 18, 16, 16, 18, 16,
         65, 65, 65, 65, 65, 48, 48, 40, 24, 18, 60, 20],
        "ProductDetailsTable",
        row_height=100,
    )
    diagnostic = _add_table_sheet(
        workbook,
        "工商候选诊断",
        "企业工商模糊搜索候选、相似度与采用决定",
        QCC_DIAGNOSTIC_HEADERS,
        qcc_diagnostic_rows,
        [24, 36, 24, 16, 12, 58],
        "QccCandidateDiagnosticsTable",
        row_height=42,
    )

    if main_rows:
        last_main_row = len(main_rows) + 2
        for row_index in range(3, last_main_row + 1):
            main.cell(row_index, 5).number_format = "0.00%"
        main.conditional_formatting.add(
            f"E3:E{last_main_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=PALE_RED,
                mid_type="num",
                mid_value=0.75,
                mid_color=PALE_AMBER,
                end_type="num",
                end_value=1,
                end_color=PALE_GREEN,
            ),
        )
        for row_index, row in enumerate(detail_rows, 3):
            for column_index in (6, 7, 8):
                detail.cell(row_index, column_index).number_format = "0"
            for column_index in (9, 10):
                cell = detail.cell(row_index, column_index)
                if cell.value == "verified":
                    cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
                elif cell.value == "needs_review":
                    cell.fill = PatternFill("solid", fgColor=PALE_AMBER)
                elif cell.value == "rejected":
                    cell.fill = PatternFill("solid", fgColor=PALE_RED)
        last_row = len(detail_rows) + 2
        for column in ("F", "G", "H"):
            detail.conditional_formatting.add(
                f"{column}3:{column}{last_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color=PALE_RED,
                    mid_type="num",
                    mid_value=70,
                    mid_color=PALE_AMBER,
                    end_type="num",
                    end_value=100,
                    end_color=PALE_GREEN,
                ),
            )
    if qcc_diagnostic_rows:
        last_diagnostic_row = len(qcc_diagnostic_rows) + 2
        for row_index in range(3, last_diagnostic_row + 1):
            diagnostic.cell(row_index, 4).number_format = "0.00%"
            decision = diagnostic.cell(row_index, 5)
            decision.fill = PatternFill(
                "solid",
                fgColor=PALE_GREEN if decision.value == "是" else PALE_AMBER,
            )
        diagnostic.conditional_formatting.add(
            f"D3:D{last_diagnostic_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=PALE_RED,
                mid_type="num",
                mid_value=0.75,
                mid_color=PALE_AMBER,
                end_type="num",
                end_value=1,
                end_color=PALE_GREEN,
            ),
        )

    target_dir = Path(output_dir).expanduser()
    if not target_dir.is_absolute():
        target_dir = Path.cwd() / target_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_suffix = f"_{run_id[:8]}" if run_id else ""
    output_path = target_dir / f"机器人产品检索核验结果_{timestamp}{run_suffix}.xlsx"
    workbook.save(output_path)
    return output_path.resolve()
