from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from rapidfuzz import fuzz, process


PRODUCT_NAME_HEADER = "产品名称"


class ProductInventoryWorkbookError(ValueError):
    """Raised when the selected product inventory workbook cannot be used."""


@dataclass(frozen=True)
class ProductNameSimilarity:
    query_name: str
    matched_name: str
    score: float
    explanation: str


def resolve_inventory_workbook_path(path: str | Path) -> Path:
    raw_path = str(path or "").strip()
    if not raw_path:
        raise ProductInventoryWorkbookError("请选择已有产品库存 Excel 文件")
    workbook_path = Path(raw_path).expanduser()
    if not workbook_path.is_absolute():
        workbook_path = Path.cwd() / workbook_path
    workbook_path = workbook_path.resolve()
    if workbook_path.suffix.lower() != ".xlsx":
        raise ProductInventoryWorkbookError("已有产品库存文件必须是 .xlsx 格式")
    if not workbook_path.is_file():
        raise ProductInventoryWorkbookError(f"已有产品库存文件不存在：{workbook_path}")
    return workbook_path


def _name_key(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalized)


@lru_cache(maxsize=8)
def _load_product_names_cached(
    workbook_path: str,
    modified_time_ns: int,
) -> tuple[str, ...]:
    del modified_time_ns  # The value forms part of the cache key.
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ProductInventoryWorkbookError(
            f"已有产品库存文件无法读取：{workbook_path}（{exc}）"
        ) from exc

    names_by_key: dict[str, str] = {}
    try:
        for worksheet in workbook.worksheets:
            header_row = 0
            product_name_column = 0
            for row_index, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row)),
                start=1,
            ):
                for column_index, cell in enumerate(row, start=1):
                    if str(cell.value or "").strip() == PRODUCT_NAME_HEADER:
                        header_row = row_index
                        product_name_column = column_index
                        break
                if product_name_column:
                    break
            if not product_name_column:
                continue

            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                min_col=product_name_column,
                max_col=product_name_column,
                values_only=True,
            ):
                display_name = str(row[0] or "").strip()
                key = _name_key(display_name)
                if display_name and key and key not in names_by_key:
                    names_by_key[key] = display_name
    finally:
        workbook.close()

    if not names_by_key:
        raise ProductInventoryWorkbookError(
            f"已有产品库存文件中未找到“{PRODUCT_NAME_HEADER}”列或有效产品名称："
            f"{workbook_path}"
        )
    return tuple(names_by_key.values())


def load_inventory_product_names(path: str | Path) -> tuple[str, ...]:
    workbook_path = resolve_inventory_workbook_path(path)
    return _load_product_names_cached(
        str(workbook_path),
        workbook_path.stat().st_mtime_ns,
    )


def _similarity_conclusion(query_key: str, matched_key: str, score: float) -> str:
    if query_key and query_key == matched_key:
        return "名称标准化后完全一致"
    if score >= 90:
        return "名称高度相似"
    if score >= 75:
        return "名称较相似"
    if score >= 50:
        return "名称部分相似"
    return "名称差异较大"


def compare_product_name(
    product_name: str,
    inventory_product_names: Iterable[str],
    *,
    inventory_filename: str = "",
) -> ProductNameSimilarity:
    query_name = str(product_name or "").strip()
    query_key = _name_key(query_name)
    choices_by_key: dict[str, str] = {}
    for raw_name in inventory_product_names:
        display_name = str(raw_name or "").strip()
        key = _name_key(display_name)
        if display_name and key and key not in choices_by_key:
            choices_by_key[key] = display_name

    if not query_key or not choices_by_key:
        explanation = "未获得可用于名称对比的产品名称"
        if inventory_filename:
            explanation += f"；库存文件：{inventory_filename}"
        return ProductNameSimilarity(query_name, "", 0.0, explanation)

    extracted = process.extractOne(
        query_key,
        tuple(choices_by_key),
        scorer=fuzz.WRatio,
        processor=None,
    )
    if extracted is None:
        return ProductNameSimilarity(
            query_name,
            "",
            0.0,
            "未在库存表的产品名称列中找到可对比名称",
        )

    matched_key, raw_score, _ = extracted
    matched_name = choices_by_key[matched_key]
    score = round(float(raw_score), 2)
    conclusion = _similarity_conclusion(query_key, matched_key, score)
    file_note = f"库存文件：{inventory_filename}；" if inventory_filename else ""
    explanation = (
        f"{file_note}最相似已有产品：{matched_name}；{conclusion}。"
        "仅比较产品名称；名称经全/半角、大小写、空格和标点统一后，"
        "使用 RapidFuzz WRatio 计算，未使用库存表其他字段。"
    )
    return ProductNameSimilarity(query_name, matched_name, score, explanation)


def compare_product_names_from_workbook(
    product_names: Iterable[str],
    inventory_workbook_path: str | Path,
) -> list[ProductNameSimilarity]:
    workbook_path = resolve_inventory_workbook_path(inventory_workbook_path)
    inventory_names = load_inventory_product_names(workbook_path)
    return [
        compare_product_name(
            product_name,
            inventory_names,
            inventory_filename=workbook_path.name,
        )
        for product_name in product_names
    ]
