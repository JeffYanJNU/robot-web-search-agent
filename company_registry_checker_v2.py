#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""企业名称批量匹配 V2。

改进点：
1. 将名称拆成品牌主体、地区、行业词、组织形式，不再只比较整串编辑距离。
2. 简称/品牌名可高置信匹配完整工商名称。
3. 地区冲突会显著降分，避免把不同地区的同品牌主体当成同一法人。
4. 保留 Top3、歧义、匹配原因和模糊度，便于人工复核。
5. 可选 OpenCC 统一繁简体；未安装时使用常见字符映射兜底。
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
import unicodedata
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz, process
except ImportError:  # pragma: no cover
    fuzz = None
    process = None
    from difflib import SequenceMatcher

try:  # 可选：pip install opencc-python-reimplemented
    from opencc import OpenCC

    _OPENCC = OpenCC("t2s")
except ImportError:  # pragma: no cover
    _OPENCC = None


SOURCE_NAME_HEADERS = ("企查查匹配企业名称", "企业名称", "公司名称", "公司名")
QUERY_HEADERS = ("公司名", "企业名称", "公司名称")
ALIAS_HEADERS = ("企业简称", "公司简称", "简称", "品牌名", "英文名称", "英文名", "曾用名", "别名")
RESULT_HEADERS = (
    "是否包含",
    "相似度",
    "模糊度",
    "最相似企业名称",
    "是否相似",
    "匹配结论",
    "是否歧义",
    "匹配原因",
    "候选Top3",
    "匹配来源",
)

BRACKET_PATTERN = re.compile(r"\(([^()]*)\)")

# 只从名称尾部移除，不能使用 str.replace，否则会误删名称主体中的“公司”等字样。
LEGAL_SUFFIXES = tuple(
    sorted(
        {
            "集团股份有限责任公司",
            "集团股份有限公司",
            "股份有限责任公司",
            "特殊普通合伙企业",
            "有限合伙企业",
            "普通合伙企业",
            "有限责任公司",
            "股份有限公司",
            "集团有限公司",
            "股份公司",
            "有限合伙",
            "普通合伙",
            "有限公司",
            "集团公司",
            "公司",
        },
        key=len,
        reverse=True,
    )
)

# 仅把尾部通用描述作为行业词。不要加入“创新、未来、华夏”等可能属于品牌的词。
INDUSTRY_SUFFIXES = tuple(
    sorted(
        {
            "人工智能技术",
            "智能制造技术",
            "机器人技术",
            "自动化技术",
            "信息技术",
            "网络技术",
            "电子技术",
            "软件技术",
            "机器人科技",
            "智能科技",
            "数字科技",
            "网络科技",
            "电子科技",
            "软件科技",
            "自动化设备",
            "智能装备",
            "智能制造",
            "人工智能",
            "机器人",
            "自动化",
            "科技",
            "技术",
            "智能",
            "信息",
            "网络",
            "电子",
            "软件",
            "动力",
            "实业",
            "控股",
            "产业",
            "发展",
        },
        key=len,
        reverse=True,
    )
)

# 省级行政区和高频城市。括号内带“省/市/区/县/州/盟/旗”的文本也会被识别为地区。
REGION_ALIASES = {
    # 直辖市/特别行政区
    "北京市": "北京", "北京": "北京", "上海市": "上海", "上海": "上海",
    "天津市": "天津", "天津": "天津", "重庆市": "重庆", "重庆": "重庆",
    "香港特别行政区": "香港", "香港": "香港", "澳门特别行政区": "澳门", "澳门": "澳门",
    # 省/自治区
    "河北省": "河北", "河北": "河北", "山西省": "山西", "山西": "山西",
    "辽宁省": "辽宁", "辽宁": "辽宁", "吉林省": "吉林", "吉林": "吉林",
    "黑龙江省": "黑龙江", "黑龙江": "黑龙江", "江苏省": "江苏", "江苏": "江苏",
    "浙江省": "浙江", "浙江": "浙江", "安徽省": "安徽", "安徽": "安徽",
    "福建省": "福建", "福建": "福建", "江西省": "江西", "江西": "江西",
    "山东省": "山东", "山东": "山东", "河南省": "河南", "河南": "河南",
    "湖北省": "湖北", "湖北": "湖北", "湖南省": "湖南", "湖南": "湖南",
    "广东省": "广东", "广东": "广东", "海南省": "海南", "海南": "海南",
    "四川省": "四川", "四川": "四川", "贵州省": "贵州", "贵州": "贵州",
    "云南省": "云南", "云南": "云南", "陕西省": "陕西", "陕西": "陕西",
    "甘肃省": "甘肃", "甘肃": "甘肃", "青海省": "青海", "青海": "青海",
    "台湾省": "台湾", "台湾": "台湾", "台灣": "台湾",
    "内蒙古自治区": "内蒙古", "内蒙古": "内蒙古",
    "广西壮族自治区": "广西", "广西": "广西",
    "西藏自治区": "西藏", "西藏": "西藏",
    "宁夏回族自治区": "宁夏", "宁夏": "宁夏",
    "新疆维吾尔自治区": "新疆", "新疆": "新疆",
    # 高频城市
    "深圳市": "深圳", "深圳": "深圳", "广州市": "广州", "广州": "广州",
    "惠州市": "惠州", "惠州": "惠州", "东莞市": "东莞", "东莞": "东莞",
    "珠海市": "珠海", "珠海": "珠海", "佛山市": "佛山", "佛山": "佛山",
    "武汉市": "武汉", "武汉": "武汉", "南京市": "南京", "南京": "南京",
    "苏州市": "苏州", "苏州": "苏州", "无锡市": "无锡", "无锡": "无锡",
    "常州市": "常州", "常州": "常州", "杭州市": "杭州", "杭州": "杭州",
    "宁波市": "宁波", "宁波": "宁波", "温州市": "温州", "温州": "温州",
    "台州市": "台州", "台州": "台州", "合肥市": "合肥", "合肥": "合肥",
    "成都市": "成都", "成都": "成都", "西安市": "西安", "西安": "西安",
    "长沙市": "长沙", "长沙": "长沙", "郑州市": "郑州", "郑州": "郑州",
    "济南市": "济南", "济南": "济南", "青岛市": "青岛", "青岛": "青岛",
    "厦门市": "厦门", "厦门": "厦门", "福州市": "福州", "福州": "福州",
    "南昌市": "南昌", "南昌": "南昌", "昆明市": "昆明", "昆明": "昆明",
    "贵阳市": "贵阳", "贵阳": "贵阳", "沈阳市": "沈阳", "沈阳": "沈阳",
    "大连市": "大连", "大连": "大连", "长春市": "长春", "长春": "长春",
    "哈尔滨市": "哈尔滨", "哈尔滨": "哈尔滨", "石家庄市": "石家庄", "石家庄": "石家庄",
}
REGION_TERMS = tuple(sorted(REGION_ALIASES, key=len, reverse=True))
ADMIN_ENDINGS = ("特别行政区", "自治区", "自治州", "地区", "省", "市", "区", "县", "州", "盟", "旗")

COMMON_TRADITIONAL_MAP = str.maketrans(
    {
        "術": "术", "臺": "台", "灣": "湾", "數": "数", "華": "华", "國": "国",
        "機": "机", "器": "器", "網": "网", "絡": "络", "電": "电", "腦": "脑",
        "資": "资", "訊": "讯", "實": "实", "業": "业", "創": "创", "聯": "联",
        "動": "动", "產": "产", "發": "发", "展": "展", "責": "责", "團": "团",
    }
)


@dataclass(frozen=True)
class CompanyRecord:
    name: str
    normalized: str
    sheet: str
    row: int


@dataclass(frozen=True)
class CompanyProfile:
    record: CompanyRecord
    full: str
    no_legal: str
    core: str
    regions: frozenset[str]
    industries: tuple[str, ...]
    legal_suffix: str
    has_explicit_region: bool


@dataclass(frozen=True)
class MatchCandidate:
    profile: CompanyProfile
    score: float
    conclusion: str
    reason: str
    location_conflict: bool


def normalize_unicode(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold().strip()
    if _OPENCC is not None:
        text = _OPENCC.convert(text)
    else:
        text = text.translate(COMMON_TRADITIONAL_MAP)
    return text


def normalize_company_name(value: object) -> str:
    text = normalize_unicode(value)
    return re.sub(r"[^0-9a-z\u3400-\u9fff]+", "", text)


def query_cache_key(value: object) -> str:
    return normalize_company_name(value)


def strip_one_legal_suffix(text: str) -> tuple[str, str]:
    for suffix in LEGAL_SUFFIXES:
        suffix_key = normalize_company_name(suffix)
        if text.endswith(suffix_key) and len(text) > len(suffix_key):
            return text[: -len(suffix_key)], suffix_key
    return text, ""


def is_region_fragment(text: str) -> bool:
    value = normalize_company_name(text)
    return bool(
        value in REGION_ALIASES
        or any(value.endswith(ending) for ending in ADMIN_ENDINGS)
    )


def canonical_region(text: str) -> str:
    value = normalize_company_name(text)
    if value in REGION_ALIASES:
        return REGION_ALIASES[value]
    for ending in ADMIN_ENDINGS:
        if value.endswith(ending) and len(value) > len(ending):
            return value[: -len(ending)]
    return value


def extract_leading_region(text: str) -> tuple[str, str]:
    for term in REGION_TERMS:
        key = normalize_company_name(term)
        if text.startswith(key) and len(text) > len(key):
            return text[len(key):], REGION_ALIASES[term]
    return text, ""


def strip_industry_suffixes(text: str) -> tuple[str, tuple[str, ...]]:
    industries: list[str] = []
    current = text
    while current:
        matched = False
        for suffix in INDUSTRY_SUFFIXES:
            key = normalize_company_name(suffix)
            if current.endswith(key) and len(current) > len(key):
                current = current[: -len(key)]
                industries.append(key)
                matched = True
                break
        if not matched:
            break
    return current or text, tuple(industries)


def build_profile(record: CompanyRecord) -> CompanyProfile:
    raw = normalize_unicode(record.name)
    bracket_values = [normalize_company_name(x) for x in BRACKET_PATTERN.findall(raw)]
    bracket_regions = {canonical_region(x) for x in bracket_values if is_region_fragment(x)}

    # 地区括号不进入主体；非地区括号仍保留，防止误把品牌括号内容删除。
    without_region_brackets = BRACKET_PATTERN.sub(
        lambda m: "" if is_region_fragment(m.group(1)) else m.group(1),
        raw,
    )
    full = normalize_company_name(raw)
    no_bracket_region = normalize_company_name(without_region_brackets)
    no_legal, legal_suffix = strip_one_legal_suffix(no_bracket_region)

    no_region_prefix, leading_region = extract_leading_region(no_legal)
    regions = set(bracket_regions)
    if leading_region:
        regions.add(leading_region)

    # 兼容“市柔宇科技股份有限公司（深圳）”这种地区被挪到末尾的写法。
    if bracket_regions and no_region_prefix.startswith("市") and len(no_region_prefix) > 1:
        no_region_prefix = no_region_prefix[1:]

    core, industries = strip_industry_suffixes(no_region_prefix)
    core = core or no_region_prefix or no_legal or full

    return CompanyProfile(
        record=record,
        full=full,
        no_legal=no_legal,
        core=core,
        regions=frozenset(regions),
        industries=industries,
        legal_suffix=legal_suffix,
        has_explicit_region=bool(regions),
    )


def ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if fuzz is not None:
        return float(fuzz.ratio(a, b))
    return SequenceMatcher(None, a, b).ratio() * 100.0


def partial_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if fuzz is not None:
        return float(fuzz.partial_ratio(a, b))
    shorter, longer = sorted((a, b), key=len)
    if shorter in longer:
        return 100.0
    return SequenceMatcher(None, shorter, longer).ratio() * 100.0


def weighted_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if fuzz is not None:
        return float(fuzz.WRatio(a, b))
    return max(ratio(a, b), partial_ratio(a, b) * 0.9)


def ngram_dice(a: str, b: str, n: int = 2) -> float:
    if not a or not b:
        return 0.0
    if len(a) < n or len(b) < n:
        return 100.0 if a == b else 0.0
    ga = {a[i:i+n] for i in range(len(a) - n + 1)}
    gb = {b[i:i+n] for i in range(len(b) - n + 1)}
    return 200.0 * len(ga & gb) / (len(ga) + len(gb)) if ga or gb else 0.0


def score_profiles(query: CompanyProfile, candidate: CompanyProfile) -> MatchCandidate:
    if query.full == candidate.full:
        return MatchCandidate(candidate, 100.0, "精确入库", "完整企业名称标准化后相同", False)

    q_regions = set(query.regions)
    c_regions = set(candidate.regions)
    location_conflict = bool(q_regions and c_regions and q_regions.isdisjoint(c_regions))
    location_same = bool(q_regions and c_regions and not q_regions.isdisjoint(c_regions))

    # 地区从前缀移动到括号等结构变化，不改变企业主体、地区、行业和组织形式。
    if (
        query.core == candidate.core
        and q_regions
        and q_regions == c_regions
        and query.industries == candidate.industries
        and query.legal_suffix == candidate.legal_suffix
    ):
        return MatchCandidate(
            candidate, 100.0, "结构等价入库",
            "企业主体、地区、行业和组织形式一致，仅地区位置或括号结构不同",
            False,
        )

    if query.no_legal == candidate.no_legal:
        score = 98.0
        reason = "仅组织形式存在差异"
        conclusion = "高置信同名"
    elif query.core == candidate.core and len(query.core) >= 2:
        score = 96.0
        if not query.legal_suffix:
            reason = "品牌主体完全一致，查询名称可能是简称"
            conclusion = "高置信简称"
        else:
            reason = "品牌主体完全一致"
            conclusion = "高置信同主体"
    else:
        core_ratio = ratio(query.core, candidate.core)
        core_partial = partial_ratio(query.core, candidate.core)
        base_ratio = weighted_ratio(query.no_legal, candidate.no_legal)
        dice = ngram_dice(query.core, candidate.core)

        score = 0.50 * core_ratio + 0.20 * core_partial + 0.15 * base_ratio + 0.15 * dice

        shorter = min(len(query.core), len(candidate.core))
        longer = max(len(query.core), len(candidate.core))
        contained = query.core in candidate.core or candidate.core in query.core
        coverage = shorter / longer if longer else 0.0
        if contained and shorter >= 3:
            score = max(score, 84.0 + 10.0 * coverage)

        if core_ratio >= 88:
            conclusion = "可能同主体"
            reason = "品牌主体高度相似，可能存在少量错字或增删字"
        elif contained and shorter >= 3:
            conclusion = "可能简称或扩展名"
            reason = "一个品牌主体包含另一个品牌主体"
        elif core_ratio >= 70:
            conclusion = "需人工复核"
            reason = "品牌主体存在一定相似，但不足以自动确认"
        else:
            conclusion = "低置信"
            reason = "相似主要来自地区、行业或组织形式，品牌主体差异较大"

    if location_same:
        score += 2.0
        reason += "；地区一致"

    if location_conflict:
        # 同品牌不同地区通常是不同法人，必须显著降分。
        score -= 28.0
        score = min(score, 72.0)
        conclusion = "地域冲突"
        reason += f"；查询地区{sorted(q_regions)}与候选地区{sorted(c_regions)}冲突"

    # 极短主体只允许完全一致高分，避免两个字近似造成大量误报。
    if min(len(query.core), len(candidate.core)) <= 2 and query.core != candidate.core:
        score = min(score, 68.0)
        conclusion = "低置信"
        reason += "；品牌主体过短且不完全一致"

    return MatchCandidate(candidate, round(max(0.0, min(score, 100.0)), 2), conclusion, reason, location_conflict)


class CompanyMatcher:
    def __init__(
        self,
        records: Iterable[CompanyRecord],
        alias_pairs: Iterable[tuple[str, str]] = (),
    ):
        self.profiles = [build_profile(record) for record in records]
        if not self.profiles:
            raise ValueError("入库文件中没有可用的企业名称。")

        self.by_full: dict[str, list[CompanyProfile]] = defaultdict(list)
        self.by_no_legal: dict[str, list[CompanyProfile]] = defaultdict(list)
        self.by_core: dict[str, list[CompanyProfile]] = defaultdict(list)
        self.by_record_name: dict[str, list[CompanyProfile]] = defaultdict(list)
        for profile in self.profiles:
            self.by_full[profile.full].append(profile)
            self.by_no_legal[profile.no_legal].append(profile)
            self.by_core[profile.core].append(profile)
            self.by_record_name[profile.record.normalized].append(profile)

        self.alias_index: dict[str, list[CompanyProfile]] = defaultdict(list)
        for alias, target_name in alias_pairs:
            alias_key = normalize_company_name(alias)
            target_key = normalize_company_name(target_name)
            if not alias_key or not target_key:
                continue
            for profile in self.by_record_name.get(target_key, ()):
                if profile not in self.alias_index[alias_key]:
                    self.alias_index[alias_key].append(profile)

        self.core_choices = list(self.by_core)
        self.base_choices = list(self.by_no_legal)

    def _candidate_pool(self, query: CompanyProfile, limit: int = 40) -> set[CompanyProfile]:
        pool: set[CompanyProfile] = set()
        pool.update(self.by_full.get(query.full, ()))
        pool.update(self.by_no_legal.get(query.no_legal, ()))
        pool.update(self.by_core.get(query.core, ()))

        if process is not None:
            for key, _score, _idx in process.extract(
                query.core, self.core_choices, scorer=fuzz.WRatio, limit=limit
            ):
                pool.update(self.by_core[key])
            for key, _score, _idx in process.extract(
                query.no_legal, self.base_choices, scorer=fuzz.WRatio, limit=max(10, limit // 2)
            ):
                pool.update(self.by_no_legal[key])
        else:  # 小数据兜底
            pool.update(self.profiles)

        # 简称包含关系补召回。5千条规模直接扫描开销可接受。
        if len(query.core) >= 2:
            for profile in self.profiles:
                if query.core in profile.core or profile.core in query.core:
                    pool.add(profile)
        return pool

    def match(self, query_value: object, top_k: int = 3) -> tuple[list[MatchCandidate], bool]:
        query_record = CompanyRecord(
            str(query_value or "").strip(), normalize_company_name(query_value), "", 0
        )
        query = build_profile(query_record)
        candidates = [score_profiles(query, item) for item in self._candidate_pool(query)]

        # 英文名、品牌名、曾用名等不能依赖字符相似度，优先使用明确别名映射。
        for profile in self.alias_index.get(query.full, ()):
            candidates.append(
                MatchCandidate(
                    profile=profile,
                    score=99.0,
                    conclusion="别名映射",
                    reason="查询名称命中企业简称、英文名、品牌名或曾用名映射",
                    location_conflict=False,
                )
            )

        # 同一企业可能同时从算法和别名进入候选，只保留最高分。
        best_by_name: dict[str, MatchCandidate] = {}
        for item in candidates:
            key = item.profile.record.normalized
            current = best_by_name.get(key)
            if current is None or item.score > current.score:
                best_by_name[key] = item
        candidates = list(best_by_name.values())
        candidates.sort(
            key=lambda x: (x.score, len(x.profile.core), x.profile.record.name),
            reverse=True,
        )
        top = candidates[:top_k]

        ambiguous = False
        if top and top[0].score >= 80.0:
            near_top = [x for x in candidates if x.score >= 80.0 and top[0].score - x.score <= 2.0]
            distinct_names = {x.profile.record.normalized for x in near_top}
            ambiguous = len(distinct_names) > 1

            # 简称没有地区，而库中同主体存在多个地区法人时，不能自动唯一确认。
            same_core = self.by_core.get(query.core, [])
            if not query.has_explicit_region and len({x.record.normalized for x in same_core}) > 1:
                ambiguous = True
        return top, ambiguous


def find_header(ws, candidates: Sequence[str], max_scan_rows: int = 10):
    candidate_set = {normalize_company_name(item) for item in candidates}
    max_row = min(ws.max_row, max_scan_rows)
    for row, values in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1
    ):
        for col, value in enumerate(values, start=1):
            if normalize_company_name(value) in candidate_set:
                return row, col, str(value).strip()
    return None


def split_aliases(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[；;、|\n]+", text) if item.strip()]


def load_source_data(source_path: Path) -> tuple[list[CompanyRecord], list[tuple[str, str]]]:
    """读取标准企业名称，并自动吸收同一行的简称/英文名/品牌名/曾用名字段。"""
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise ValueError("入库文件中没有工作表。")
        ws = workbook.worksheets[0]
        header = find_header(ws, SOURCE_NAME_HEADERS)
        if header:
            header_row, name_col, _ = header
        else:
            header_row, name_col = 1, 2

        alias_keys = {normalize_company_name(x) for x in ALIAS_HEADERS}
        header_values = next(
            ws.iter_rows(
                min_row=header_row, max_row=header_row, values_only=True
            ),
            (),
        )
        alias_cols = [
            col
            for col, value in enumerate(header_values, start=1)
            if normalize_company_name(value) in alias_keys
        ]

        records: list[CompanyRecord] = []
        alias_pairs: list[tuple[str, str]] = []
        for row, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            value = values[name_col - 1] if name_col <= len(values) else None
            normalized = normalize_company_name(value)
            if not normalized:
                continue
            record = CompanyRecord(str(value).strip(), normalized, ws.title, row)
            records.append(record)
            for col in alias_cols:
                alias_value = values[col - 1] if col <= len(values) else None
                for alias in split_aliases(alias_value):
                    alias_pairs.append((alias, record.name))
        return records, alias_pairs
    finally:
        workbook.close()


def load_external_alias_pairs(path: Path | None) -> list[tuple[str, str]]:
    """读取可选别名表，字段为“别名”和“标准企业名称”。支持 xlsx/csv。"""
    if path is None:
        return []
    if not path.exists():
        raise FileNotFoundError(f"别名文件不存在：{path}")

    pairs: list[tuple[str, str]] = []
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                alias = row.get("别名") or row.get("alias")
                target = row.get("标准企业名称") or row.get("企业名称") or row.get("company_name")
                if alias and target:
                    pairs.append((alias, target))
        return pairs

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook.worksheets[0]
        alias_header = find_header(ws, ("别名", "alias"))
        target_header = find_header(ws, ("标准企业名称", "企业名称", "company_name"))
        if not alias_header or not target_header:
            raise ValueError("别名文件需要包含‘别名’和‘标准企业名称’两列。")
        header_row = max(alias_header[0], target_header[0])
        alias_col, target_col = alias_header[1], target_header[1]
        for row in range(header_row + 1, ws.max_row + 1):
            alias = ws.cell(row, alias_col).value
            target = ws.cell(row, target_col).value
            if alias and target:
                for item in split_aliases(alias):
                    pairs.append((item, str(target).strip()))
        return pairs
    finally:
        workbook.close()


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def ensure_result_columns(ws, header_row: int, query_col: int):
    existing = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value:
            existing[normalize_company_name(value)] = col

    columns = {}
    next_col = ws.max_column + 1
    for header in RESULT_HEADERS:
        key = normalize_company_name(header)
        if key in existing:
            columns[header] = existing[key]
            continue
        preferred = {"是否包含": query_col + 1, "相似度": query_col + 2}.get(header)
        col = preferred if preferred and not ws.cell(header_row, preferred).value else next_col
        while ws.cell(header_row, col).value:
            col += 1
        columns[header] = col
        ws.cell(header_row, col).value = header
        copy_cell_style(ws.cell(header_row, query_col), ws.cell(header_row, col))
        next_col = max(next_col, col + 1)
    return columns


def source_label(record: CompanyRecord) -> str:
    return f"{record.sheet}!第{record.row}行"


def process_template(
    template_path: Path,
    output_path: Path,
    matcher: CompanyMatcher,
    threshold: float,
    template_sheet: str | None = None,
):
    if template_path.resolve() == output_path.resolve():
        raise ValueError("输出文件不能覆盖模板文件，请指定其他输出路径。")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path)
    try:
        candidate_sheets = [workbook[template_sheet]] if template_sheet else workbook.worksheets
        located = None
        for ws in candidate_sheets:
            header = find_header(ws, QUERY_HEADERS)
            if header:
                located = (ws, *header)
                break
        if not located:
            raise ValueError("模板中未找到公司名称输入列。")

        ws, header_row, query_col, _ = located
        result_cols = ensure_result_columns(ws, header_row, query_col)
        template_style_row = header_row + 1 if ws.max_row > header_row else header_row

        ws.cell(header_row, result_cols["相似度"]).comment = Comment(
            "综合品牌主体、地区、行业和组织形式评分。简称主体一致可高分；明确地区冲突会降分。",
            "CompanyMatcherV2",
        )
        ws.cell(header_row, result_cols["模糊度"]).comment = Comment(
            "模糊度=1-相似度，越高表示越不确定。该值是规则分数，不是统计概率。",
            "CompanyMatcherV2",
        )

        cache = {}
        counts = {"total": 0, "exact": 0, "similar": 0, "ambiguous": 0, "blank": 0}

        for row in range(header_row + 1, ws.max_row + 1):
            query_value = ws.cell(row, query_col).value
            normalized = normalize_company_name(query_value)
            for col in result_cols.values():
                copy_cell_style(
                    ws.cell(template_style_row, min(col, ws.max_column)),
                    ws.cell(row, col),
                )

            if not normalized:
                for col in result_cols.values():
                    ws.cell(row, col).value = None
                counts["blank"] += 1
                continue

            counts["total"] += 1
            cache_key = query_cache_key(query_value)
            if cache_key not in cache:
                cache[cache_key] = matcher.match(query_value, top_k=3)
            top, ambiguous = cache[cache_key]
            best = top[0] if top else None

            score = best.score if best else 0.0
            is_exact = bool(best and best.conclusion in {"精确入库", "结构等价入库"})
            is_similar = bool(best and score >= threshold and not best.location_conflict)
            if is_exact:
                counts["exact"] += 1
            elif is_similar:
                counts["similar"] += 1
            if ambiguous:
                counts["ambiguous"] += 1

            ws.cell(row, result_cols["是否包含"]).value = "是" if is_exact else "否"
            ws.cell(row, result_cols["相似度"]).value = score / 100.0
            ws.cell(row, result_cols["相似度"]).number_format = "0.00%"
            ws.cell(row, result_cols["模糊度"]).value = (100.0 - score) / 100.0
            ws.cell(row, result_cols["模糊度"]).number_format = "0.00%"
            ws.cell(row, result_cols["最相似企业名称"]).value = best.profile.record.name if best else ""
            ws.cell(row, result_cols["是否相似"]).value = "是" if is_similar else "否"
            ws.cell(row, result_cols["匹配结论"]).value = best.conclusion if best else "无候选"
            ws.cell(row, result_cols["是否歧义"]).value = "是" if ambiguous else "否"
            ws.cell(row, result_cols["匹配原因"]).value = best.reason if best else "未找到候选"
            ws.cell(row, result_cols["候选Top3"]).value = "\n".join(
                f"{idx}. {item.profile.record.name}（{item.score:.2f}%）"
                for idx, item in enumerate(top, start=1)
            )
            ws.cell(row, result_cols["匹配来源"]).value = (
                source_label(best.profile.record) if best else ""
            )

            processed = row - header_row
            total_rows = ws.max_row - header_row
            if processed % 10 == 0 or processed == total_rows:
                print(f"匹配进度：{processed}/{total_rows}", flush=True)

        widths = {
            "是否包含": 12, "相似度": 12, "模糊度": 12, "最相似企业名称": 38,
            "是否相似": 12, "匹配结论": 18, "是否歧义": 12, "匹配原因": 52,
            "候选Top3": 60, "匹配来源": 44,
        }
        for header, width in widths.items():
            col = result_cols[header]
            letter = get_column_letter(col)
            current = ws.column_dimensions[letter].width
            if not current or current < width:
                ws.column_dimensions[letter].width = width

        ws.freeze_panes = ws.freeze_panes or ws.cell(header_row + 1, 1).coordinate
        last_col = max(query_col, *result_cols.values())
        ws.auto_filter.ref = (
            f"{get_column_letter(query_col)}{header_row}:"
            f"{get_column_letter(last_col)}{ws.max_row}"
        )

        green = PatternFill("solid", fgColor="E2F0D9")
        yellow = PatternFill("solid", fgColor="FFF2CC")
        red = PatternFill("solid", fgColor="FCE4D6")
        for row in range(header_row + 1, ws.max_row + 1):
            for header in ("是否包含", "是否相似", "是否歧义"):
                cell = ws.cell(row, result_cols[header])
                if cell.value == "是":
                    cell.fill = yellow if header == "是否歧义" else green
                elif cell.value == "否":
                    cell.fill = red if header != "是否歧义" else green
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for header in ("相似度", "模糊度", "匹配结论"):
                ws.cell(row, result_cols[header]).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            for header in ("匹配原因", "候选Top3"):
                ws.cell(row, result_cols[header]).alignment = Alignment(
                    vertical="top", wrap_text=True
                )

        workbook.save(output_path)
        return ws.title, counts
    except Exception:
        if output_path.exists():
            output_path.unlink()
        raise
    finally:
        workbook.close()


def find_default_source(base_dir: Path) -> Path:
    matches = sorted(
        (
            path for path in base_dir.glob("已入库企业信息*.xlsx")
            if not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("当前目录未找到‘已入库企业信息*.xlsx’。")
    return matches[0]


def parse_args(argv: Sequence[str] | None = None):
    parser = argparse.ArgumentParser(description="企业名称批量匹配 V2")
    parser.add_argument("--source", type=Path, help="入库企业信息 xlsx 文件")
    parser.add_argument("--template", type=Path, help="查询模板 xlsx 文件")
    parser.add_argument("--output", type=Path, help="输出结果 xlsx 文件")
    parser.add_argument(
        "--threshold", type=float, default=85.0,
        help="判断为相似企业的阈值，默认 85；地区冲突不会判为相似",
    )
    parser.add_argument("--sheet", help="模板工作表名称")
    parser.add_argument("--alias-file", type=Path, help="可选别名表：别名、标准企业名称")
    return parser.parse_args(argv)


def application_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if not 0 <= args.threshold <= 100:
        raise ValueError("--threshold 必须在 0 到 100 之间。")

    base_dir = application_dir()
    source = (args.source or find_default_source(base_dir)).resolve()
    template = (args.template or (base_dir / "查询企业.xlsx")).resolve()
    output = (args.output or (base_dir / "企业入库批量检索结果_V2.xlsx")).resolve()

    for label, path in (("入库文件", source), ("模板文件", template)):
        if not path.exists():
            raise FileNotFoundError(f"{label}不存在：{path}")

    print(f"正在读取入库文件：{source.name}", flush=True)
    records, source_aliases = load_source_data(source)
    external_aliases = load_external_alias_pairs(
        args.alias_file.resolve() if args.alias_file else None
    )
    alias_pairs = source_aliases + external_aliases
    print(f"读取完成，共 {len(records):,} 条记录；正在建立匹配索引...", flush=True)
    matcher = CompanyMatcher(records, alias_pairs=alias_pairs)
    print(
        f"索引完成，已载入 {len(records):,} 条企业记录、"
        f"{len(alias_pairs):,} 条别名映射。",
        flush=True,
    )

    sheet, counts = process_template(
        template, output, matcher, args.threshold, args.sheet
    )
    print(f"处理工作表：{sheet}")
    print(
        f"完成：查询 {counts['total']} 家，精确 {counts['exact']} 家，"
        f"相似 {counts['similar']} 家，歧义 {counts['ambiguous']} 家，"
        f"空白 {counts['blank']} 行。"
    )
    print(f"结果文件：{output}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
