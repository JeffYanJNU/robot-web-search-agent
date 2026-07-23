import json
from datetime import date, datetime, timezone

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app import main as main_module
from app.database import Base
from app.models import ProductCompanyRelation, ProductSource, RobotCompany, RobotProduct
from app.schemas import RunResult
from app.services.result_exporter import _group_products_for_export, export_run_results


def _inventory_workbook(tmp_path, *product_names: str):
    path = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["企业名称", "产品名称", "产品品类"])
    for index, product_name in enumerate(product_names, start=1):
        worksheet.append([f"企业 {index}", product_name, "机器人"])
    workbook.save(path)
    return path


def test_export_group_collapses_obvious_product_name_variants():
    first = RobotProduct(
        product_id=101,
        canonical_name="Unitree B2-W",
        original_name="Unitree B2-W",
        normalized_name="unitreeb2w",
        identity_key="unitreeb2w|b2w",
        model_number="B2-W",
        series_name="Unitree",
        authenticity_score=90,
        novelty_score=80,
    )
    package = RobotProduct(
        product_id=102,
        canonical_name="Unitree B2-W Kit",
        original_name="Unitree B2-W Kit",
        normalized_name="unitreeb2wkit",
        identity_key="unitreeb2wkit|b2w",
        model_number="B2-W",
        series_name="Unitree",
        authenticity_score=70,
        novelty_score=60,
    )
    first.sources = [
        ProductSource(
            source_url="https://example.com/b2w",
            canonical_url="https://example.com/b2w",
            content_hash="a" * 64,
            raw_content="source one",
            evidence_json="[]",
        )
    ]
    package.sources = [
        ProductSource(
            source_url="https://example.com/b2w-kit",
            canonical_url="https://example.com/b2w-kit",
            content_hash="b" * 64,
            raw_content="source two",
            evidence_json="[]",
        )
    ]

    grouped = _group_products_for_export([first, package])

    assert len(grouped) == 1
    assert grouped[0].canonical_name == "Unitree B2-W"
    assert {source.source_url for source in grouped[0].sources} == {
        "https://example.com/b2w",
        "https://example.com/b2w-kit",
    }


def test_exporter_writes_one_row_for_duplicate_product_records(tmp_path):
    inventory_path = _inventory_workbook(tmp_path, "Unitree B2-W", "Walker S2")
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        company = RobotCompany(
            canonical_name="Unitree Robotics",
            original_name="Unitree Robotics",
            country="China",
            region_type="mainland_china",
        )
        db.add(company)
        products = [
            RobotProduct(
                canonical_name=name,
                original_name=name,
                normalized_name=normalized,
                identity_key=identity,
                model_number="B2-W",
                series_name="Unitree",
                authenticity_score=score,
                novelty_score=score,
                verification_status="needs_review",
            )
            for name, normalized, identity, score in [
                ("Unitree B2-W", "unitreeb2w", "unitreeb2w|b2w", 90),
                ("Unitree B2-W Kit", "unitreeb2wkit", "unitreeb2wkit|b2w", 70),
            ]
        ]
        db.add_all(products)
        db.flush()
        for index, product in enumerate(products):
            db.add_all([
                ProductSource(
                    product_id=product.product_id,
                    source_url=f"https://example.com/b2w-{index}",
                    canonical_url=f"https://example.com/b2w-{index}",
                    content_hash=str(index) * 64,
                    raw_content=f"source {index}",
                    evidence_json="[]",
                ),
                ProductCompanyRelation(
                    product_id=product.product_id,
                    company_id=company.company_id,
                    relation_type="developer",
                    relation_score=80,
                    verification_status="verified",
                    evidence_json="[]",
                    is_primary=True,
                ),
            ])
        db.commit()

        path = export_run_results(
            db,
            RunResult(product_ids=[product.product_id for product in products]),
            pipeline_mode="product",
            lookback_days=30,
            output_dir=str(tmp_path),
            run_id="duplicate-test",
            inventory_workbook_path=str(inventory_path),
        )

    workbook = load_workbook(path, data_only=False)
    assert workbook.worksheets[0].max_row == 3
    assert workbook.worksheets[1].max_row == 3
    assert workbook.worksheets[1].cell(3, 5).value == 2


def test_exporter_writes_six_column_main_sheet_and_detail_sheet(tmp_path, monkeypatch):
    inventory_path = _inventory_workbook(tmp_path, "Walker-S2", "Unitree B2-W")
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    now = datetime.now(timezone.utc)
    with Session(engine) as db:
        company = RobotCompany(
            canonical_name="优必选",
            original_name="深圳市优必选科技股份有限公司",
            chinese_name="优必选",
            baseline_company_name="深圳市优必选科技股份有限公司",
            country="中国",
            region_type="mainland_china",
            official_website="https://ubtrobot.com",
            official_domain="ubtrobot.com",
            unified_social_credit_code="91440300TEST000001",
        )
        product = RobotProduct(
            canonical_name="Walker S2",
            original_name="Walker S2",
            normalized_name="walkers2",
            identity_key="walkers2|s2",
            model_number="S2",
            series_name="Walker",
            robot_category="人形机器人",
            launch_date=date.today(),
            launch_status="released",
            addition_type="new_product",
            authenticity_score=95,
            novelty_score=90,
            verification_status="verified",
            verification_reason="产品名称、发布事件和来源完整",
        )
        db.add_all([company, product])
        db.flush()
        quote = "优必选正式发布 Walker S2 工业人形机器人。"
        source_urls = [
            "https://ubtrobot.com/news/walker-s2",
            "https://people.com.cn/robot/walker-s2",
        ]
        for index, url in enumerate(source_urls):
            db.add(
                ProductSource(
                    product_id=product.product_id,
                    source_url=url,
                    canonical_url=url,
                    source_title="Walker S2 发布",
                    source_type="official" if index == 0 else "authority",
                    published_at=now,
                    content_hash=str(index) * 64,
                    claim_fingerprint=str(index + 2) * 64,
                    raw_content=quote,
                    evidence_json=json.dumps(
                        [
                            {
                                "evidence_type": "product_launch",
                                "quote": quote,
                                "value": "Walker S2",
                                "evidence_date": date.today().isoformat(),
                            }
                        ],
                        ensure_ascii=False,
                    ),
                )
            )
        db.add(
            ProductCompanyRelation(
                product_id=product.product_id,
                company_id=company.company_id,
                relation_type="developer",
                relation_score=100,
                verification_status="verified",
                verification_reason="官网和权威来源共同确认",
                evidence_json=json.dumps(
                    [
                        {"quote": quote, "source_url": source_urls[0]},
                        {"quote": quote, "source_url": source_urls[1]},
                    ],
                    ensure_ascii=False,
                ),
                is_primary=True,
            )
        )
        db.commit()

        result = RunResult(
            product_ids=[product.product_id],
            company_ids=[company.company_id],
            products_created=1,
            relations_created=1,
            relations_verified=1,
            queries=3,
            results=8,
            fetched=2,
        )
        path = export_run_results(
            db,
            result,
            pipeline_mode="product",
            lookback_days=30,
            output_dir=str(tmp_path),
            run_id="test-run",
            inventory_workbook_path=str(inventory_path),
        )

    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["结果主表", "详细信息"]

    main = workbook["结果主表"]
    assert main.max_column == 6
    assert [main.cell(2, column).value for column in range(1, 7)] == [
        "A｜机器人产品名称",
        "B｜关联企业（简称 / 全称）",
        "C｜产品是否存在及依据",
        "D｜产品与企业是否对应及依据",
        "E｜与已有产品名称相似度",
        "F｜相似度说明",
    ]
    assert main["A3"].value == "Walker S2"
    assert "简称：优必选" in main["B3"].value
    assert "全称：深圳市优必选科技股份有限公司" in main["B3"].value
    assert "产品存在" in main["C3"].value
    assert "明确归属" in main["D3"].value
    assert main["E3"].value == 1
    assert main["E3"].number_format == "0.00%"
    assert "最相似已有产品：Walker-S2" in main["F3"].value
    assert "仅比较产品名称" in main["F3"].value
    assert main.freeze_panes == "A3"
    assert len(main.tables) == 1
    assert main.auto_filter.ref is None

    detail = workbook["详细信息"]
    headers = {detail.cell(2, column).value: column for column in range(1, detail.max_column + 1)}
    assert detail.cell(3, headers["检索热度（来源页数）"]).value == 2
    assert detail.cell(3, headers["产品真实性评分"]).value == 95
    assert detail.cell(3, headers["新产品置信度"]).value == 90
    assert detail.cell(3, headers["关系置信度"]).value == 100
    assert "https://ubtrobot.com/news/walker-s2" in detail.cell(
        3, headers["全部产品来源"]
    ).value
    assert "权重 25" in detail.cell(3, headers["产品评分依据"]).value
    assert "Excel 企业基线" == detail.cell(3, headers["企业全称来源"]).value
    assert len(detail.tables) == 1

    monkeypatch.setattr(main_module.settings, "output_dir", str(tmp_path))
    files = main_module.list_output_files()
    assert files[0]["filename"] == path.name
    response = main_module.download_output_file(path.name)
    assert response.path == path
